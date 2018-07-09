
import boto3
import xlsx2csv
from io import BytesIO
from openpyxl import load_workbook
s3 = boto3.client('s3')
bucket = 'oncoscape-datasets'
key = 'b13f39d3-b7de-4f02-ac11-48742e832202/source.xlsx'
data = s3.get_object(Bucket=bucket, Key=key)
wb = load_workbook(filename=BytesIO(data['Body'].read()), read_only=True)
sheetNames = wb.sheetnames
print(sheetNames)


converter = xlsx2csv.Xlsx2csv(file, sheetid=0, outputencoding='utf-8')
del wb
for idx, sheetName in enumerate(sheetNames):
    converter.convert(sheetName + '.csv', sheetid=idx)



file = 'broad.xlsx'
wb = load_workbook(filename=file, read_only=True)
sheetNames = wb.sheetnames
converter = xlsx2csv.Xlsx2csv(file, sheetid=0, outputencoding='utf-8')
del wb


for idx, sheetName in enumerate(sheetNames):
    converter.convert(sheetName+'.csv', sheetid=idx)

s3.get_object
