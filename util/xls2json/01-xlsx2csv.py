import xlsx2csv
from openpyxl import load_workbook
wb = load_workbook(filename='broad.xlsx', read_only=True)
sheetNames = wb.sheetnames
del wb
converter = xlsx2csv.Xlsx2csv('broad.xlsx', sheetid=0, outputencoding='utf-8')
for idx, sheetname in enumerate(sheetNames):
    converter.convert(sheetname+'.csv', sheetid=idx)